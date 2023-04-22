from django.shortcuts import render
import pandas as pd
import numpy as np
import openpyxl
from django.http import HttpResponse
import sklearn as sk
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_absolute_percentage_error
import pickle
import requests
from django.http import JsonResponse
from bs4 import BeautifulSoup
import json
import random
# Create your views here.
def index(request):
    return render(request, 'dashboard/index.html')


def charts(request):
    return render(request, 'dashboard/charts.html')


def widgets(request):
    return render(request, 'dashboard/widgets.html')




def tables(request):
    return render(request, "dashboard/tables.html")




def grid(request):
    return render(request, "dashboard/grid.html")




def form_basic(request):
    return render(request, "dashboard/form_basic.html")


def ml_page_view(request):
    return render(request, 'ml_page.html')


def form_wizard(request):
    return render(request, "dashboard/form_wizard.html")




def buttons(request):
    return render(request, "dashboard/buttons.html")




def icon_material(request):
    return render(request, "dashboard/icon-material.html")




def icon_fontawesome(request):
    return render(request, "dashboard/icon-fontawesome.html")




def elements(request):
    return render(request, "dashboard/elements.html")




def gallery(request):
    return render(request, "dashboard/gallery.html")





def invoice(request):
    return render(request, "dashboard/invoice.html")



def chat(request):
    return render(request, "dashboard/chat.html")

from django.shortcuts import render

def read_excel(request):
    df = pd.read_excel('dashboard/Profiles Talep Final.xlsx')
    df = df.iloc[:, 1:6]
    year = 2021
    while(year < 2081):
        temp_df=df.iloc[:,-5:]
        df[year]=temp_df.mean(axis=1)
        year+=1
    populations=[70586256.0 ,
    71517100.0 ,
    72561312.0 ,
    73722988.0 ,
    74724269.0 ,
    75627384.0 ,
    76667864.0 ,
    77695904.0 ,
    78741053.0 ,
    79814871.0 ,
    80810525.0 ,
    82003882.0 ,
    83154997.0 ,
    83614362.0 ,
    84680273.0 ,
    85679953.83359557 ,
    86673605.9268678 ,
    87649178.78034161 ,
    88605961.31182502 ,
    89543083.86679992 ,
    90459627.92254707 ,
    91354730.80571996 ,
    92227582.7004141 ,
    93077541.33884321 ,
    93900523.73293717 ,
    94696114.00433126 ,
    95463878.32307665 ,
    96203626.20291995 ,
    96915384.57123727 ,
    97599320.9761468 ,
    98255721.64568375 ,
    98884772.07955034 ,
    99486604.92000681 ,
    100061364.77703613 ,
    100611303.57713124 ,
    101136738.4649444 ,
    101637859.92672901 ,
    102114752.73385437 ,
    102567362.03425798 ,
    102995788.55627279 ,
    103400164.94215903 ,
    103780358.54965647 ,
    104136230.75264356 ,
    104467670.84968504 ,
    104782035.00101666 ,
    105079166.62893273 ,
    105358757.56457305 ,
    105620556.48576039 ,
    105864267.19135728 ,
    106089848.7916444 ,
    106297283.33503762 ,
    106486287.58543245 ,
    106656559.35783006 ,
    106807934.09604296 ,
    106940522.50150637 ,
    107054500.10082029 ,
    107150114.226921 ,
    107227726.90342104 ,
    107287885.65363511 ,
    107331548.89269152 ,
    107359860.53542086 ,
    107373839.83287533 ,
    107374487.08721276 ,
    107362977.1298808 ,
    107340540.6416998 ,
    107308400.3250083 ,
    107267622.30443914 ,
    107219251.76111041 ,
    107164391.72013338 ,
    107103902.86023411 ,
    107038212.03018986 ,
    106967653.3289267 ,
    106892489.0490344 ,
    106812826.9000175 ]
    elecons_reg=np.zeros(74)
    elecons_exp=np.zeros(74)
    smo_val=np.zeros(74)
    smo_tre=np.zeros(74)
    FTs=np.zeros(74)
    Demands=np.zeros(74)
    known=[2.198,
    2.264,
    2.162,
    2.334,
    2.49,
    2.577,
    2.583,
    2.669,
    2.76,
    2.897,
    3.082,
    3.149,
    3.094,
    3.142 ]
    known_demands=[2155148590.7,
    161914714.4,
    156877556.5,
    172069454,
    186063429.8,
    194891768.6,
    198033092.7,
    207370367.8,
    217325306.3,
    231223681.3,
    249058038.1,
    258230224.4,
    257281560.7,
    262716325.4]
    for i in range(14):
        elecons_reg[i]=known[i]
    for i in range(14):
        elecons_exp[i]=known[i]
    for i in range(14):
        Demands[i]=known_demands[i]
    m=8.16977E-08
    b=-3.638143043
    for i in range(14,74):
        elecons_reg[i]=m*populations[i]+b
    a=1
    be=0.383264
    smo_val[0]=155148590.7
    Demands[0]=155148590.7
    elecons_exp[0]=2.198
    FTs[1]=155148590.7
    for i in range(1,74):
        smo_val[i]=a*Demands[i]+(1-a)*(smo_val[i-1]+smo_tre[i-1])
        smo_tre[i]=be*(smo_val[i]-smo_val[i-1])+(1-be)*smo_tre[i-1]
        if(i<73):
            FTs[i+1]=smo_val[i]+smo_tre[i]
        if(i>=13 and i<73):
            Demands[i+1]=FTs[i+1]
        if(i>=14):
            elecons_exp[i]=Demands[i]/populations[i]
    demands_simp_reg=pd.DataFrame()
    demands_simp_reg["populations"]=populations
    demands_simp_reg["elecons"]=elecons_reg
    demands_simp_reg["Demands"]=demands_simp_reg["populations"]*demands_simp_reg["elecons"]
    demands_exp=pd.DataFrame()
    demands_exp["populations"]=populations
    demands_exp["elecons"]=elecons_exp
    demands_exp["Demands"]=demands_exp["populations"]*demands_exp["elecons"]
    expo_yillik = {}
    year=2007
    for i in range(74):
        expo_yillik[year]= demands_exp["Demands"][i]
        year+=1
    regre_yillik = {}
    year=2007
    for i in range(74):
        regre_yillik[year]= demands_simp_reg["Demands"][i]
        year+=1
    expo_ort_saatlik = {}
    for year in range(2007,2081):
        expo_ort_saatlik[year]= expo_yillik[year]/8760
    regre_ort_saatlik = {}
    for year in range(2007,2081):
        regre_ort_saatlik[year]= regre_yillik[year]/8760
    movav_regre_yillik_tahmin_df = pd.DataFrame()
    movav_expo_yillik_tahmin_df = pd.DataFrame()
    for i in range(2016,2081):
        movav_regre_yillik_tahmin_df[i]=df[i]*regre_ort_saatlik[i]
    for i in range(2016,2081):
        movav_expo_yillik_tahmin_df[i]=df[i]*expo_ort_saatlik[i]
    context = {'data': movav_expo_yillik_tahmin_df.to_html(index=False, classes='table table-striped',
                            border=1, header=True, na_rep='-',
                            float_format=lambda x: f'{x:,.2f}')}
    return render(request, 'excel_data.html', context)

def chart_view(request):
    df = pd.read_excel('dashboard/Profiles Talep Final.xlsx')
    df = df.iloc[:, 1:6]
    year = 2021
    while(year < 2081):
        temp_df=df.iloc[:,-5:]
        df[year]=temp_df.mean(axis=1)
        year+=1
    populations=[70586256.0 ,
    71517100.0 ,
    72561312.0 ,
    73722988.0 ,
    74724269.0 ,
    75627384.0 ,
    76667864.0 ,
    77695904.0 ,
    78741053.0 ,
    79814871.0 ,
    80810525.0 ,
    82003882.0 ,
    83154997.0 ,
    83614362.0 ,
    84680273.0 ,
    85679953.83359557 ,
    86673605.9268678 ,
    87649178.78034161 ,
    88605961.31182502 ,
    89543083.86679992 ,
    90459627.92254707 ,
    91354730.80571996 ,
    92227582.7004141 ,
    93077541.33884321 ,
    93900523.73293717 ,
    94696114.00433126 ,
    95463878.32307665 ,
    96203626.20291995 ,
    96915384.57123727 ,
    97599320.9761468 ,
    98255721.64568375 ,
    98884772.07955034 ,
    99486604.92000681 ,
    100061364.77703613 ,
    100611303.57713124 ,
    101136738.4649444 ,
    101637859.92672901 ,
    102114752.73385437 ,
    102567362.03425798 ,
    102995788.55627279 ,
    103400164.94215903 ,
    103780358.54965647 ,
    104136230.75264356 ,
    104467670.84968504 ,
    104782035.00101666 ,
    105079166.62893273 ,
    105358757.56457305 ,
    105620556.48576039 ,
    105864267.19135728 ,
    106089848.7916444 ,
    106297283.33503762 ,
    106486287.58543245 ,
    106656559.35783006 ,
    106807934.09604296 ,
    106940522.50150637 ,
    107054500.10082029 ,
    107150114.226921 ,
    107227726.90342104 ,
    107287885.65363511 ,
    107331548.89269152 ,
    107359860.53542086 ,
    107373839.83287533 ,
    107374487.08721276 ,
    107362977.1298808 ,
    107340540.6416998 ,
    107308400.3250083 ,
    107267622.30443914 ,
    107219251.76111041 ,
    107164391.72013338 ,
    107103902.86023411 ,
    107038212.03018986 ,
    106967653.3289267 ,
    106892489.0490344 ,
    106812826.9000175 ]
    elecons_reg=np.zeros(74)
    elecons_exp=np.zeros(74)
    smo_val=np.zeros(74)
    smo_tre=np.zeros(74)
    FTs=np.zeros(74)
    Demands=np.zeros(74)
    known=[2.198,
    2.264,
    2.162,
    2.334,
    2.49,
    2.577,
    2.583,
    2.669,
    2.76,
    2.897,
    3.082,
    3.149,
    3.094,
    3.142 ]
    known_demands=[2155148590.7,
    161914714.4,
    156877556.5,
    172069454,
    186063429.8,
    194891768.6,
    198033092.7,
    207370367.8,
    217325306.3,
    231223681.3,
    249058038.1,
    258230224.4,
    257281560.7,
    262716325.4]
    for i in range(14):
        elecons_reg[i]=known[i]
    for i in range(14):
        elecons_exp[i]=known[i]
    for i in range(14):
        Demands[i]=known_demands[i]
    m=8.16977E-08
    b=-3.638143043
    for i in range(14,74):
        elecons_reg[i]=m*populations[i]+b
    a=1
    be=0.383264
    smo_val[0]=155148590.7
    Demands[0]=155148590.7
    elecons_exp[0]=2.198
    FTs[1]=155148590.7
    for i in range(1,74):
        smo_val[i]=a*Demands[i]+(1-a)*(smo_val[i-1]+smo_tre[i-1])
        smo_tre[i]=be*(smo_val[i]-smo_val[i-1])+(1-be)*smo_tre[i-1]
        if(i<73):
            FTs[i+1]=smo_val[i]+smo_tre[i]
        if(i>=13 and i<73):
            Demands[i+1]=FTs[i+1]
        if(i>=14):
            elecons_exp[i]=Demands[i]/populations[i]
    demands_simp_reg=pd.DataFrame()
    demands_simp_reg["populations"]=populations
    demands_simp_reg["elecons"]=elecons_reg
    demands_simp_reg["Demands"]=demands_simp_reg["populations"]*demands_simp_reg["elecons"]
    demands_exp=pd.DataFrame()
    demands_exp["populations"]=populations
    demands_exp["elecons"]=elecons_exp
    demands_exp["Demands"]=demands_exp["populations"]*demands_exp["elecons"]
    expo_yillik = {}
    year=2007
    for i in range(74):
        expo_yillik[year]= demands_exp["Demands"][i]
        year+=1
    regre_yillik = {}
    year=2007
    for i in range(74):
        regre_yillik[year]= demands_simp_reg["Demands"][i]
        year+=1
    expo_ort_saatlik = {}
    for year in range(2007,2081):
        expo_ort_saatlik[year]= expo_yillik[year]/8760
    regre_ort_saatlik = {}
    for year in range(2007,2081):
        regre_ort_saatlik[year]= regre_yillik[year]/8760
    movav_regre_yillik_tahmin_df = pd.DataFrame()
    movav_expo_yillik_tahmin_df = pd.DataFrame()
    for i in range(2016,2081):
        movav_regre_yillik_tahmin_df[i]=df[i]*regre_ort_saatlik[i]
    for i in range(2016,2081):
        movav_expo_yillik_tahmin_df[i]=df[i]*expo_ort_saatlik[i]
    y_data = movav_expo_yillik_tahmin_df[2024].tolist()
    return render(request, 'chart.html', {'y_data': y_data})

def show_terminals(request):
    #dataframe1 = openpyxl.load_workbook('dashboard/Terminaller.xlsx')
    df = pd.read_excel('dashboard/Terminaller.xlsx', usecols="D")
    print(df)
    to_list=df.values.tolist()
    context = {'data': df}

    if request.method == 'POST':
        terminals = request.POST.getlist('terminals')
        filename = 'dashboard/Terminaller.xlsx'
        wb = openpyxl.load_workbook(filename=filename)
        ws = wb['Sheet1']
        last_row = ws.max_row + 1
        for i, tag in enumerate(terminals):
            ws.cell(row=last_row+i, column=1, value=tag)
        wb.save(filename)
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        with open(filename, 'rb') as f:
            response.write(f.read())
        #name = request.POST['name']
        #email = request.POST['email']
        #message = request.POST['message']
        #terminal=request.POST['terminal']

        # Load existing workbook
        #wb = load_workbook('dashboard/Terminaller.xlsx')
        #ws = wb.active

        # Find first empty row
        #row = ws.max_row + 1

        # Add data to cells
        #ws.cell(row, 1, name)
        #ws.cell(row, 2, email)
        #ws.cell(row, 3, message)
        #ws.cell(row, 4, terminal)
        # Save workbook
        #wb.save('dashboard/Terminaller.xlsx')

        # Return response
        return render(request, 'show_terminal.html', {
            "option": to_list
    })
     # Render HTML form
    return render(request, 'show_terminal.html', {
            "option": to_list
    })
def lineer_reg_ml(request):
    filename = 'dashboard/minmax_scaler_model.pkl'
    scaler = pickle.load(open(filename, 'rb'))
    with open('dashboard/lineer_reg_model.pkl', 'rb') as f:lr_loaded = pickle.load(f)
    r = requests.get("https://seffaflik.epias.com.tr/transparency/piyasalar/gop/ptf.xhtml")
    soup = BeautifulSoup(r.content,"html.parser")
    ptf_values=[]
    for tr in soup.find_all('tr')[1:]:
        td_list = tr.find_all('td')
        ptf_value = td_list[3].text.strip() # PTF (USD/MWh) sütunu
        ptv_val_temp=ptf_value.replace(',', '.')
        ptf_values.append(float(ptv_val_temp))
    ptf_values.pop(0)
    tuketim=pd.read_excel("dashboard/GercekZamanliTuketim.xlsx")
    tuketim_list=tuketim["tuketim"].to_list()
    columns=['1 Hafta önceki tüketim',
            '1 Saat Önceki PTF',
            '2 Saat Önceki PTF',
            '3 Saat Önceki PTF',
            '4 Saat Önceki PTF',
            '11 Saat Önceki PTF']
    count=0
    for i in range(24):
        temp=pd.DataFrame(columns=columns)
        list2=[]
        list2.append(tuketim_list[-168-count])
        list2.append(ptf_values[-1+count])
        list2.append(ptf_values[-2+count])
        list2.append(ptf_values[-3+count])
        list2.append(ptf_values[-4+count])
        list2.append(ptf_values[-11+count])
        temp.loc[len(temp)] = list2
        temp=scaler.transform(temp)
        ptf=lr_loaded.predict(temp)
        ptf_values.append(ptf[0])
    PTFs=ptf_values[-24:]
    context = {'PTFs': PTFs}
    return render(request, 'lineer_reg_ml.html', context)

def gradientb_reg_ml(request):
    filename = 'dashboard/minmax_scaler_model.pkl'
    scaler = pickle.load(open(filename, 'rb'))
    with open('dashboard/gb_reg_model.pkl', 'rb') as f:lr_loaded = pickle.load(f)
    r = requests.get("https://seffaflik.epias.com.tr/transparency/piyasalar/gop/ptf.xhtml")
    soup = BeautifulSoup(r.content,"html.parser")
    ptf_values=[]
    for tr in soup.find_all('tr')[1:]:
        td_list = tr.find_all('td')
        ptf_value = td_list[3].text.strip() # PTF (USD/MWh) sütunu
        ptv_val_temp=ptf_value.replace(',', '.')
        ptf_values.append(float(ptv_val_temp))
    ptf_values.pop(0)
    tuketim=pd.read_excel("dashboard/GercekZamanliTuketim.xlsx")
    tuketim_list=tuketim["tuketim"].to_list()
    columns=['1 Hafta önceki tüketim',
            '1 Saat Önceki PTF',
            '2 Saat Önceki PTF',
            '3 Saat Önceki PTF',
            '4 Saat Önceki PTF',
            '11 Saat Önceki PTF']
    count=0
    for i in range(24):
        temp=pd.DataFrame(columns=columns)
        list2=[]
        list2.append(tuketim_list[-168-count])
        list2.append(ptf_values[-1+count])
        list2.append(ptf_values[-2+count])
        list2.append(ptf_values[-3+count])
        list2.append(ptf_values[-4+count])
        list2.append(ptf_values[-11+count])
        temp.loc[len(temp)] = list2
        temp=scaler.transform(temp)
        ptf=lr_loaded.predict(temp)
        ptf_values.append(ptf[0])
    PTFs=ptf_values[-24:]
    print(PTFs)
    context = {'PTFs': PTFs}
    return render(request, 'gb_reg_ml.html', context)

def randomf_reg_ml(request):
    filename = 'dashboard/minmax_scaler_model.pkl'
    scaler = pickle.load(open(filename, 'rb'))
    with open('dashboard/rf_reg_model.pkl', 'rb') as f:lr_loaded = pickle.load(f)
    r = requests.get("https://seffaflik.epias.com.tr/transparency/piyasalar/gop/ptf.xhtml")
    soup = BeautifulSoup(r.content,"html.parser")
    ptf_values=[]
    for tr in soup.find_all('tr')[1:]:
        td_list = tr.find_all('td')
        ptf_value = td_list[3].text.strip() # PTF (USD/MWh) sütunu
        ptv_val_temp=ptf_value.replace(',', '.')
        ptf_values.append(float(ptv_val_temp))
    ptf_values.pop(0)
    tuketim=pd.read_excel("dashboard/GercekZamanliTuketim.xlsx")
    tuketim_list=tuketim["tuketim"].to_list()
    columns=['1 Hafta önceki tüketim',
            '1 Saat Önceki PTF',
            '2 Saat Önceki PTF',
            '3 Saat Önceki PTF',
            '4 Saat Önceki PTF',
            '11 Saat Önceki PTF']
    count=0
    for i in range(24):
        temp=pd.DataFrame(columns=columns)
        list2=[]
        list2.append(tuketim_list[-168-count])
        list2.append(ptf_values[-1+count])
        list2.append(ptf_values[-2+count])
        list2.append(ptf_values[-3+count])
        list2.append(ptf_values[-4+count])
        list2.append(ptf_values[-11+count])
        temp.loc[len(temp)] = list2
        temp=scaler.transform(temp)
        ptf=lr_loaded.predict(temp)
        ptf_values.append(ptf[0])
    PTFs=ptf_values[-24:]
    context = {'PTFs': PTFs}
    return render(request, 'rf_reg_ml.html', context)